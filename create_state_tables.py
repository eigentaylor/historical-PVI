import pandas as pd
import os
import utils
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

# Load the data
df = pd.read_csv('presidential_margins.csv')

# Ensure output directory exists
output_dir = 'state_tables'
os.makedirs(output_dir, exist_ok=True)

# Get unique states
states = df['abbr'].unique()

# Create a dictionary to store data for each state
state_data = {}
min_year = df['year'].min()
max_year = df['year'].max()

for state in states:
    state_df = df[df['abbr'] == state]
    years = state_df['year']
    pres_margin = state_df['pres_margin']
    national_margin = state_df['national_margin']
    relative_margin = state_df['relative_margin']

    # Format the data using utils.lean_str
    formatted_data = {
        'Year': years,
        'Relative Margin': [utils.lean_str(m) for m in relative_margin],
        'Presidential Margin': [utils.lean_str(m) for m in pres_margin],
        'National Margin': [utils.lean_str(m) for m in national_margin],
        'Relative Margin (Float)': relative_margin,
        'Presidential Margin (Float)': pres_margin,
        'National Margin (Float)': national_margin
    }

    # Convert to DataFrame
    state_table = pd.DataFrame(formatted_data)

    # Save to dictionary
    state_data[state] = state_table

# Combine all state tables into a single Excel file with plots
output_file = os.path.join(output_dir, f'state_margins_{min_year}_{max_year}.xlsx')
# delete file if it exists
if os.path.exists(output_file):
    os.remove(output_file)
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    for state, table in state_data.items():
        # Save the table to a sheet
        table.to_excel(writer, sheet_name=state, index=False)

        # Reuse plots from the state_trends directory
        plot_path = os.path.join('state_trends', f'{state}_trend_subplot.png')
        if os.path.exists(plot_path):
            img = Image(plot_path)
            worksheet = writer.sheets[state]

            # Calculate the row to place the image below the table
            max_row = worksheet.max_row + 2  # Add some spacing
            img.anchor = f'A{max_row}'
            worksheet.add_image(img)
        else:
            print(f"Warning: Plot for {state} not found in state_trends directory.")

        # Adjust column widths
        for col_num, column_cells in enumerate(worksheet.columns, 1):
            max_length = 0
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width

print(f"Excel file with plots created at {output_file}")

# Ensure rankings directory exists
rankings_dir = os.path.join('state_trends', 'rankings')
os.makedirs(rankings_dir, exist_ok=True)

# Generate rankings for each year
years = df['year'].unique()

for year in years:
    year_df = df[df['year'] == year]
    rankings = year_df[['abbr', 'relative_margin', 'pres_margin', 'national_margin']].sort_values(by='relative_margin', ascending=False)

    # Write rankings to a text file
    rankings_file = os.path.join(rankings_dir, f'{year}_rankings.txt')
    with open(rankings_file, 'w') as f:
        f.write(f"Rankings for {year}\n")
        f.write("====================\n")
        i = 1
        for idx, row in rankings.iterrows():
            f.write(f"{i}.\t{row['abbr']}:\t Rel {utils.lean_str(row['relative_margin'])}\t Actual {utils.lean_str(row['pres_margin'])}\t Nat {utils.lean_str(row['national_margin'])}\n")
            i += 1

print(f"Rankings created in {rankings_dir}")
